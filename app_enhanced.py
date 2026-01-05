from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

import pandas as pd
import os
from pathlib import Path
from io import BytesIO
from datetime import datetime
from typing import Optional, Dict, Any, List
from functools import lru_cache


app = FastAPI(title="Accessories Sales Dashboard", version="1.0.0")

# -------------------- CORS --------------------
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -------------------- PATHS --------------------
BASE_DIR = Path(__file__).resolve().parent
TEMPLATES_DIR = BASE_DIR / "templates"
STATIC_DIR = BASE_DIR / "static"

if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# -------------------- REQUEST MODEL --------------------
class FilterRequest(BaseModel):
    quarter: Optional[str] = ""
    month: Optional[str] = ""
    location: Optional[str] = ""
    model: Optional[str] = ""


# -------------------- HELPERS --------------------
def _clean_col_name(c: str) -> str:
    return " ".join(str(c).replace("\t", " ").strip().split())


# If your CSV headers have extra spaces, this normalization will fix it
CANON_RENAME = {
    "No of Billied Ros": "No of Billied Ros",
    "Acc Sale throughROs (GNDP)In Rs": "Acc Sale throughROs (GNDP)In Rs",
    "Acc Sale throughROs (MRP) In Rs": "Acc Sale throughROs (MRP) In Rs",
    "No of Counter ROs": "No of Counter ROs",
    "Acc Sale throughCounter (GNDP) In Rs": "Acc Sale throughCounter (GNDP) In Rs",
    "Acc Sale throughCounter (MRP) In Rs": "Acc Sale throughCounter (MRP) In Rs",
    "Acc Revenue (GNDP) / RO": "Acc Revenue (GNDP) / RO",
    "Acc Revenue (MRP) / RO": "Acc Revenue (MRP) / RO",
}

REQUIRED_COLS = ["Fiscal Quarter", "Fiscal Month", "Location", "Model Group"]

NUMERIC_COLS = [
    "No of Billied Ros",
    "Acc Sale throughROs (GNDP)In Rs",
    "Acc Sale throughROs (MRP) In Rs",
    "No of Counter ROs",
    "Acc Sale throughCounter (GNDP) In Rs",
    "Acc Sale throughCounter (MRP) In Rs",
    "Acc Revenue (GNDP) / RO",
    "Acc Revenue (MRP) / RO",
]

TABLE_COLS = [
    "Fiscal Quarter", "Fiscal Month", "Location", "Model Group",
    "No of Billied Ros",
    "Acc Sale throughROs (GNDP)In Rs",
    "Acc Sale throughROs (MRP) In Rs",
    "No of Counter ROs",
    "Acc Sale throughCounter (GNDP) In Rs",
    "Acc Sale throughCounter (MRP) In Rs",
    "Acc Revenue (GNDP) / RO",
    "Acc Revenue (MRP) / RO",
]

INDIAN_FINANCIAL_MONTHS = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar']

df = pd.DataFrame()
quarters: List[str] = []
months: List[str] = []
locations: List[str] = []
models: List[str] = []


def load_csv() -> pd.DataFrame:
    global quarters, months, locations, models

    possible_csv = [
        "/mnt/data/Accessories.csv",           # Render disk (optional)
        str(BASE_DIR / "Accessories.csv"),     # repo root (your GitHub file)
        "Accessories.csv",
        "./Accessories.csv",
        os.path.join(os.getcwd(), "Accessories.csv"),
    ]

    csv_path = None
    for p in possible_csv:
        if os.path.exists(p):
            csv_path = p
            break

    if not csv_path:
        print("Accessories.csv not found. Tried:")
        for p in possible_csv:
            print(" -", p)
        return pd.DataFrame()

    print("Loading Accessories.csv from:", csv_path)

    # CSV load
    _df = pd.read_csv(csv_path, low_memory=False)

    # clean headers
    _df.columns = [_clean_col_name(c) for c in _df.columns]

    # rename by canonical map
    rename_map = {}
    for c in _df.columns:
        cc = _clean_col_name(c)
        rename_map[c] = CANON_RENAME.get(cc, cc)
    _df = _df.rename(columns=rename_map)

    # required columns check
    for r in REQUIRED_COLS:
        if r not in _df.columns:
            raise ValueError(f"Missing required column in CSV: {r}")

    # ensure string columns
    for c in REQUIRED_COLS:
        _df[c] = _df[c].astype(str).str.strip()

    # numeric conversion
    for c in NUMERIC_COLS:
        if c in _df.columns:
            _df[c] = pd.to_numeric(_df[c], errors="coerce").fillna(0)

    # dropdown lists
    quarters = sorted([x for x in _df["Fiscal Quarter"].dropna().unique().tolist() if str(x).strip()])
    all_months = [x for x in _df["Fiscal Month"].dropna().unique().tolist() if str(x).strip()]
    months = sorted(all_months, key=lambda x: INDIAN_FINANCIAL_MONTHS.index(x) if x in INDIAN_FINANCIAL_MONTHS else 999)
    locations = sorted([x for x in _df["Location"].dropna().unique().tolist() if str(x).strip()])
    models = sorted([x for x in _df["Model Group"].dropna().unique().tolist() if str(x).strip()])

    print("Rows loaded:", len(_df))
    return _df


def compute_totals(_df: pd.DataFrame) -> Dict[str, Any]:
    totals: Dict[str, Any] = {}
    if _df is None or _df.empty:
        for c in NUMERIC_COLS:
            totals[c] = 0.0
        totals["No of Billied Ros"] = 0
        totals["No of Counter ROs"] = 0
        return totals

    for c in NUMERIC_COLS:
        totals[c] = float(_df[c].sum()) if c in _df.columns else 0.0

    totals["No of Billied Ros"] = int(_df["No of Billied Ros"].sum()) if "No of Billied Ros" in _df.columns else 0
    totals["No of Counter ROs"] = int(_df["No of Counter ROs"].sum()) if "No of Counter ROs" in _df.columns else 0
    return totals


# Load once at startup
try:
    df = load_csv()
except Exception as e:
    print("CSV load error:", e)
    df = pd.DataFrame()


# Cache filtering for speed
@lru_cache(maxsize=512)
def cached_filter(q: str, m: str, l: str, md: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    f = df
    if q:
        f = f[f["Fiscal Quarter"] == q]
    if m:
        f = f[f["Fiscal Month"] == m]
    if l:
        f = f[f["Location"] == l]
    if md:
        f = f[f["Model Group"] == md]
    return f


# -------------------- ROUTES --------------------
@app.get("/")
def root():
    index_file = TEMPLATES_DIR / "index.html"
    if not index_file.exists():
        raise HTTPException(status_code=500, detail="templates/index.html not found")
    return FileResponse(str(index_file), media_type="text/html")


@app.get("/api/filter-options")
def filter_options():
    return {
        "quarters": quarters,
        "months": months,
        "locations": locations,
        "models": models,
    }


@app.post("/api/get-data")
def get_data(req: FilterRequest):
    if df.empty:
        return {"data": [], "totals": compute_totals(pd.DataFrame()), "count": 0}

    q = (req.quarter or "").strip()
    m = (req.month or "").strip()
    l = (req.location or "").strip()
    md = (req.model or "").strip()

    f = cached_filter(q, m, l, md)
    totals = compute_totals(f)

    # send only required columns for faster response
    if not f.empty:
        f2 = f[TABLE_COLS].copy()
        data = f2.to_dict("records")
    else:
        data = []

    return {"data": data, "totals": totals, "count": int(len(f))}


@app.post("/api/export-excel")
def export_excel(req: FilterRequest):
    if df.empty:
        raise HTTPException(status_code=400, detail="No data loaded")

    q = (req.quarter or "").strip()
    m = (req.month or "").strip()
    l = (req.location or "").strip()
    md = (req.model or "").strip()

    f = cached_filter(q, m, l, md)
    f_export = f[TABLE_COLS].copy() if not f.empty else pd.DataFrame(columns=TABLE_COLS)
    totals = compute_totals(f)

    output = BytesIO()
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            f_export.to_excel(writer, sheet_name="Sales Data", index=False)
            ws = writer.sheets["Sales Data"]

            # add totals row
            totals_row = {c: "" for c in TABLE_COLS}
            totals_row["Fiscal Quarter"] = "TOTAL"
            for c in NUMERIC_COLS:
                if c in totals_row:
                    totals_row[c] = totals.get(c, 0)

            start_row = len(f_export) + 2
            pd.DataFrame([totals_row]).to_excel(
                writer, sheet_name="Sales Data", startrow=start_row - 1, index=False, header=False
            )

            thin = Side(style="thin", color="D1D5DB")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # header format
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            # cell borders
            for row in ws.iter_rows(min_row=2, max_row=len(f_export) + 1):
                for cell in row:
                    cell.border = border

            # totals row format
            totals_excel_row = len(f_export) + 2
            for cell in ws[totals_excel_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
                cell.border = border

            # auto width
            for col_cells in ws.columns:
                col_letter = col_cells[0].column_letter
                max_len = 0
                for c in col_cells:
                    v = "" if c.value is None else str(c.value)
                    max_len = max(max_len, len(v))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=accessories_sales_filtered.xlsx"},
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@app.get("/api/health")
def health():
    return {"status": "ok", "timestamp": datetime.now().isoformat()}
